import json
import shutil
from pathlib import Path
from openpyxl import load_workbook
from tkinter import Tk, filedialog

# ================================
# ⚙️ CONFIG
# ================================
BASE_DIR = Path(__file__).resolve().parent
MODELS_DIR = BASE_DIR / "model"

PREDICTIONS_JSON = MODELS_DIR / "predictions.json"
TEMPLATE_EXCEL = MODELS_DIR / "Template.xlsx"  # your format/template file

# ================================
# 📍 CATEGORY → CELL MAPPING
# ================================
def category_cell_map():
    """
    Decide where each category count goes.

    Sheet can be:
    - sheet name (str)
    - sheet index (int, 1-based)

    Example:
    {
        "tos": {"sheet": 2, "cell": "A2"}
    }
    """
    return {
        "TOS": {
            "sheet": "IPCR",
            "cell": "C30"
        },
        "SLM": {
            "sheet": "IPCR",
            "cell": "C21"
        },
        "Course Guide": {
            "sheet": "IPCR",
            "cell": "C20"
        },
        "Grading Sheet": {
            "sheet": "IPCR",
            "cell": "C34"
        },
        "Syllabus": {
            "sheet": "IPCR",
            "cell": "C19"
        },
        # add more categories here
    }

# ================================
# 📂 SAVE FILE DIALOG
# ================================
def prompt_save_location(default_name="classified_report.xlsx"):
    root = Tk()
    root.withdraw()  # hide main window

    file_path = filedialog.asksaveasfilename(
        title="Save Excel Report As",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel Files", "*.xlsx")]
    )

    root.destroy()
    return file_path

# ================================
# 🧠 MAIN LOGIC
# ================================
def create_excel_from_template():
    if not PREDICTIONS_JSON.exists():
        raise FileNotFoundError("❌ predictions.json not found")

    if not TEMPLATE_EXCEL.exists():
        raise FileNotFoundError("❌ Excel template not found in models folder")

    # Ask user where to save the new file
    save_path = prompt_save_location()
    if not save_path:
        print("⚠️ Save cancelled by user")
        return

    save_path = Path(save_path)

    # Duplicate the template
    shutil.copy(TEMPLATE_EXCEL, save_path)

    # Load prediction data
    with open(PREDICTIONS_JSON, "r", encoding="utf-8") as f:
        predictions = json.load(f)

    category_counts = predictions.get("category_counts", {})
    mapping = category_cell_map()

    # Open duplicated workbook
    wb = load_workbook(save_path)

    for category, count in category_counts.items():
        if category not in mapping:
            continue  # skip categories you haven't mapped

        target = mapping[category]
        sheet_ref = target["sheet"]
        cell = target["cell"]

        # Select worksheet
        if isinstance(sheet_ref, int):
            ws = wb.worksheets[sheet_ref - 1]
        else:
            ws = wb[sheet_ref]

        ws[cell] = count

    wb.save(save_path)
    print(f"✅ Excel report created at:\n{save_path}")

# ================================
# ▶️ RUN
# ================================
if __name__ == "__main__":
    create_excel_from_template()
